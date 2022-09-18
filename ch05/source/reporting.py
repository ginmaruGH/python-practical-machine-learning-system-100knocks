import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Font

from .calculation import Calc

class Reporting:

    @staticmethod
    def data_export(df, ws, row_start, col_start):
        """
        データフレームのデータを
        Excelシートの指定された行、列に貼り付ける
        スタイル定義
        """
        side = Side(style="thin", color="008080")
        border = Border(top=side, bottom=side, left=side, right=side)

        rows = dataframe_to_rows(df, index=False, header=True)

        for row_no, row in enumerate(rows, row_start):
            for col_no, value in enumerate(row, col_start):
                cell = ws.cell(row_no, col_no)
                cell.value = value
                cell.border = border
                if row_no == row_start:
                    cell.fill = PatternFill(
                        patternType="solid", fgColor="008080")
                    cell.font = Font(bold=True, color="FFFFFF")

    def make_report_headquarters(
            target_data: pd.DataFrame,
            max_str_date: str,
            m_store,
            output_folder: str
        ):
        """
        本部向けのデータを揃えたレポートを作成
        指定された年月度のレポートを作成する
        """
        rank = Calc.get_rank_df(target_data, m_store)
        cancel_rank = Calc.get_cancel_rank_df(target_data, m_store)

        # Excel出力処理
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "サマリーレポート（本部向け）"

        cell = ws.cell(1, 1)
        cell.value = f"本部向け {max_str_date}月度 サマリーリポート"
        cell.font = Font(bold=True, color="008080", size=20)

        cell = ws.cell(3, 2)
        cell.value = f"{max_str_date}月度 売上総額"
        cell.font = Font(bold=True, color="008080", size=20)

        cell = ws.cell(3, 6)
        cell.value = f"{'{:,}'.format(rank['total_amount'].sum())}"
        cell.font = Font(bold=True, color="008080", size=20)

        # 売上ランキングを出力
        cell = ws.cell(5, 2)
        cell.value = "売上ランキング"
        cell.font = Font(bold=True, color="008080", size=16)

        # 表に貼り付け位置
        Reporting.data_export(rank, ws, row_start=6, col_start=2)

        # キャンセル率ランキングを出力
        cell = ws.cell(5, 8)
        cell.value = "キャンセル率ランキング"
        cell.font = Font(bold=True, color="008080", size=16)

        # 表に貼り付け位置
        Reporting.data_export(cancel_rank, ws, row_start=6, col_start=8)

        # Workbookの保存して閉じる
        wb.save(os.path.join(output_folder, f"report_hq_{max_str_date}.xlsx"))
        wb.close()

    def make_report_headquarters_r2(
            target_data_list: list,
            m_store: pd.DataFrame,
            output_folder: str
        ):
        """
        過去月を同時に出力できるように改修した本部向けレポートを出力する
        """
        wb = openpyxl.Workbook()
        file_year_month = ""
        for tmp in target_data_list:

            df = pd.DataFrame(tmp)
            df_date = pd.to_datetime(df["order_accept_date"]).max()
            year_month = df_date.strftime("%Y%m")

            # 初回のみ、ファイル名用の年月を取得する
            if file_year_month == "":
                file_year_month = year_month

            rank = Calc.get_rank_df(df, m_store)
            cancel_rank = Calc.get_cancel_rank_df(df, m_store)

            # WorkSheet作成
            ws = wb.create_sheet(title=f"{file_year_month}月度")

            cell = ws.cell(1, 1)
            cell.value = f"本部向け {file_year_month}月度 サマリーリポート"
            cell.font = Font(bold=True, color="008080", size=20)

            cell = ws.cell(3, 2)
            cell.value = f"{file_year_month}月度 売上総額"
            cell.font = Font(bold=True, color="008080", size=20)

            cell = ws.cell(3, 6)
            cell.value = f"{'{:,}'.format(rank['total_amount'].sum())}"
            cell.font = Font(bold=True, color="008080", size=20)

            # 売上ランキングを出力
            cell = ws.cell(5, 2)
            cell.value = "売上ランキング"
            cell.font = Font(bold=True, color="008080", size=16)

            # 表に貼り付け位置
            Reporting.data_export(rank, ws, row_start=6, col_start=2)

            # キャンセル率ランキングを出力
            cell = ws.cell(5, 8)
            cell.value = "キャンセル率ランキング"
            cell.font = Font(bold=True, color="008080", size=16)

            # 表に貼り付け位置
            Reporting.data_export(cancel_rank, ws, row_start=6, col_start=8)

        # デフォルトシートの削除
        wb.remove(wb.worksheets[0])

        # ブックを保存する
        wb.save(os.path.join(output_folder, f"report_hq_{file_year_month}.xlsx"))
        wb.close()

    def make_report_store(
            target_data,
            target_id,
            m_store,
            max_str_date,
            output_folder
        ):
        """
        各店舗向けのデータを揃えたレポートを作成
        """
        rank = Calc.get_store_rank(target_id, target_data, m_store)
        sale = Calc.get_store_sale(target_id, target_data, m_store)
        cancel_rank = Calc.get_store_cancel_rank(target_id, target_data, m_store)
        cancel_count = Calc.get_store_cancel_count(target_id, target_data)
        delivery_df = Calc.get_delivery_rank_df(target_id, target_data, m_store)
        delivery_rank = Calc.get_store_delivery_rank(target_id, target_data, m_store)

        store_name = m_store.loc[
            m_store['store_id'] == target_id
        ]["store_name"].values[0]

        # Excel出力処理
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "店舗向けレポーティング"

        cell = ws.cell(1, 1)
        cell.value = f"{store_name} {max_str_date}月度 サマリーリポート"
        cell.font = Font(bold=True, color="008080", size=20)

        cell = ws.cell(3, 2)
        cell.value = f"{max_str_date}月度 売上総額"
        cell.font = Font(bold=True, color="008080", size=20)

        cell = ws.cell(3, 6)
        cell.value = f"{'{:,}'.format(sale.values[0])}"
        cell.font = Font(bold=True, color="008080", size=20)

        # 売上ランキングを出力
        cell = ws.cell(5, 2)
        cell.value = "売上ランキング"
        cell.font = Font(bold=True, color="008080", size=16)

        cell = ws.cell(5, 5)
        cell.value = f"{rank}位"
        cell.font = Font(bold=True, color="008080", size=16)

        cell = ws.cell(6, 2)
        cell.value = "売上データ"
        cell.font = Font(bold=True, color="008080", size=16)

        # 売上データ表の貼り付け
        tmp_df = target_data.loc[
            (target_data["store_id"] == target_id) &
            (target_data["status"].isin([1, 2]))
        ]
        tmp_df = tmp_df[[
            "order_accept_date", "customer_id", "total_amount",
            "takeout_name", "status_name"
        ]]
        Reporting.data_export(tmp_df, ws, row_start=7, col_start=2)

        # キャンセル率ランキングを出力
        cell = ws.cell(5, 8)
        cell.value = "キャンセル率ランキング"
        cell.font = Font(bold=True, color="008080", size=16)

        cell = ws.cell(5, 12)
        cell.value = f"{cancel_rank}位 {cancel_count.values[0]}回"
        cell.font = Font(bold=True, color="008080", size=16)

        cell = ws.cell(6, 8)
        cell.value = "キャンセルデータ"
        cell.font = Font(bold=True, color="008080", size=16)

        # キャンセルデータ表の貼り付け
        tmp_df = target_data.loc[
            (target_data["store_id"] == target_id) &
            (target_data["status"] == 9)
        ]
        tmp_df = tmp_df[[
            "order_accept_date", "customer_id", "total_amount",
            "takeout_name", "status_name"
        ]]
        Reporting.data_export(tmp_df, ws, row_start=7, col_start=8)

        # 配達完了までの時間を出力
        ave_time = delivery_df.loc[
            delivery_df["store_id"] == target_id
        ]["delta"].values[0]

        cell = ws.cell(5, 14)
        cell.value = "配達完了までの時間ランキング"
        cell.font = Font(bold=True, color="008080", size=16)

        cell = ws.cell(5, 18)
        cell.value = f"{delivery_rank}位 {ave_time}分"
        cell.font = Font(bold=True, color="008080", size=16)

        cell = ws.cell(6, 14)
        cell.value = "各店舗の配達時間ランク"
        cell.font = Font(bold=True, color="008080", size=16)

        # 各店舗の配達時間ランク表の貼り付け
        Reporting.data_export(delivery_df, ws, row_start=7, col_start=14)

        # Workbookの保存して閉じる
        wb.save(os.path.join(
            output_folder, f"{target_id}_{store_name}_report_{max_str_date}.xlsx"
        ))
        wb.close()

    def make_report_store_r2(
        target_data_list: list,
        m_store: pd.DataFrame,
        target_id,
        output_folder: str
    ):
        """
        """
        wb = openpyxl.Workbook()
        file_year_month = ""
        for tmp in target_data_list:

            df = pd.DataFrame(tmp)
            df_date = pd.to_datetime(df["order_accept_date"]).max()
            year_month = df_date.strftime("%Y%m")

            if file_year_month == "":
                file_year_month = year_month

            rank = Calc.get_store_rank(target_id, df, m_store)
            sale = Calc.get_store_sale(target_id, df, m_store)
            cancel_rank = Calc.get_store_cancel_rank(target_id, df, m_store)
            cancel_count = Calc.get_store_cancel_count(target_id, df)
            delivery_df = Calc.get_delivery_rank_df(target_id, df, m_store)
            delivery_rank = Calc.get_store_delivery_rank(target_id, df, m_store)

            store_name = m_store.loc[
                m_store['store_id'] == target_id
            ]["store_name"].values[0]

            # ワークシート作成
            ws = wb.create_sheet(title=f"{year_month}月度")

            # Excel出力処理
            cell = ws.cell(1, 1)
            cell.value = f"{store_name} {year_month}月度 サマリーリポート"
            cell.font = Font(bold=True, color="008080", size=20)

            cell = ws.cell(3, 2)
            cell.value = f"{year_month}月度 売上総額"
            cell.font = Font(bold=True, color="008080", size=20)

            cell = ws.cell(3, 6)
            cell.value = f"{'{:,}'.format(sale.values[0])}"
            cell.font = Font(bold=True, color="008080", size=20)

            # 売上ランキングを出力
            cell = ws.cell(5, 2)
            cell.value = "売上ランキング"
            cell.font = Font(bold=True, color="008080", size=16)

            cell = ws.cell(5, 5)
            cell.value = f"{rank}位"
            cell.font = Font(bold=True, color="008080", size=16)

            cell = ws.cell(6, 2)
            cell.value = "売上データ"
            cell.font = Font(bold=True, color="008080", size=16)

            # 売上データ表の貼り付け
            tmp_df = df.loc[
                (df["store_id"] == target_id) &
                (df["status"].isin([1, 2]))
            ]
            tmp_df = tmp_df[[
                "order_accept_date", "customer_id", "total_amount",
                "takeout_name", "status_name"
            ]]
            Reporting.data_export(tmp_df, ws, row_start=7, col_start=2)

            # キャンセル率ランキングを出力
            cell = ws.cell(5, 8)
            cell.value = "キャンセル率ランキング"
            cell.font = Font(bold=True, color="008080", size=16)

            cell = ws.cell(5, 12)
            cell.value = f"{cancel_rank}位 {cancel_count.values[0]}回"
            cell.font = Font(bold=True, color="008080", size=16)

            cell = ws.cell(6, 8)
            cell.value = "キャンセルデータ"
            cell.font = Font(bold=True, color="008080", size=16)

            # キャンセルデータ表の貼り付け
            tmp_df = df.loc[
                (df["store_id"] == target_id) &
                (df["status"] == 9)
            ]
            tmp_df = tmp_df[[
                "order_accept_date", "customer_id", "total_amount",
                "takeout_name", "status_name"
            ]]
            Reporting.data_export(tmp_df, ws, row_start=7, col_start=8)

            # 配達完了までの時間を出力
            ave_time = delivery_df.loc[
                delivery_df["store_id"] == target_id
            ]["delta"].values[0]

            cell = ws.cell(5, 14)
            cell.value = "配達完了までの時間ランキング"
            cell.font = Font(bold=True, color="008080", size=16)

            cell = ws.cell(5, 18)
            cell.value = f"{delivery_rank}位 {ave_time}分"
            cell.font = Font(bold=True, color="008080", size=16)

            cell = ws.cell(6, 14)
            cell.value = "各店舗の配達時間ランク"
            cell.font = Font(bold=True, color="008080", size=16)

            # 各店舗の配達時間ランク表の貼り付け
            Reporting.data_export(delivery_df, ws, row_start=7, col_start=14)

        # デフォルトシートの削除
        wb.remove(wb.worksheets[0])

        # ブックを保存する
        wb.save(os.path.join(
            output_folder, f"{target_id}_{store_name}_report_{file_year_month}.xlsx"
        ))
        wb.close()
